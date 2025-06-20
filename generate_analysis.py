#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
import sys
import os

def generate_packing_analysis(source_file, output_file=None):
    """Generate packing list analysis from source packing list"""
    
    # Read source data
    df = pd.read_excel(source_file, sheet_name='DATA', header=1)
    
    # Group by CALIBRE (which becomes SIZE) and CAT
    grouped = df.groupby(['CALIBRE', 'CAT'])['CANTIDAD - Cajas'].sum().reset_index()
    
    # Pivot to get categories as columns
    pivot = grouped.pivot(index='CALIBRE', columns='CAT', values='CANTIDAD - Cajas').fillna(0)
    
    # Ensure all category columns exist - handle both numeric and string representations
    if 1 not in pivot.columns and '1' not in pivot.columns:
        pivot[1] = 0
    if '1*' not in pivot.columns:
        pivot['1*'] = 0
    if 'II' not in pivot.columns and 2 not in pivot.columns:
        pivot['II'] = 0
    
    # Rename columns to match target format - handle numeric 1 as well
    rename_dict = {}
    if 1 in pivot.columns:
        rename_dict[1] = 'CAT1'
    if '1' in pivot.columns:
        rename_dict['1'] = 'CAT1'
    if '1*' in pivot.columns:
        rename_dict['1*'] = 'CAT1*'
    if 'II' in pivot.columns:
        rename_dict['II'] = 'CAT2'
    if 2 in pivot.columns:
        rename_dict[2] = 'CAT2'
        
    pivot = pivot.rename(columns=rename_dict)
    
    # Calculate total boxes per size
    pivot['BOXES'] = pivot['CAT1'] + pivot['CAT1*'] + pivot['CAT2']
    
    # Calculate total percentage
    total_boxes = pivot['BOXES'].sum()
    pivot['TOTAL%'] = (pivot['BOXES'] / total_boxes * 100).round(2)
    
    # Reset index to make CALIBRE a column
    result = pivot.reset_index()
    result = result.rename(columns={'CALIBRE': 'SIZE'})
    
    # Reorder columns
    result = result[['SIZE', 'BOXES', 'TOTAL%', 'CAT1', 'CAT1*', 'CAT2']]
    
    # Sort by SIZE
    result = result.sort_values('SIZE')
    
    # Convert to integers where appropriate
    for col in ['SIZE', 'BOXES', 'CAT1', 'CAT1*', 'CAT2']:
        result[col] = result[col].astype(int)
    
    # Create Excel file with formatting
    if output_file is None:
        output_file = source_file.replace('.xlsx', '_ANALYSIS.xlsx')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hoja1'
    
    # Add title
    ws.merge_cells('B8:G8')
    ws['B8'] = 'PACKING LIST- PALTAS NUÑEZ'
    ws['B8'].font = Font(bold=True, size=14)
    ws['B8'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ['SIZE', 'BOXES', 'TOTAL%', 'BOXES CAT1', 'BOXES CAT1*', 'BOXES CAT2']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for idx, row in result.iterrows():
        row_num = idx + 11
        ws.cell(row=row_num, column=2, value=row['SIZE'])
        ws.cell(row=row_num, column=3, value=row['BOXES'])
        ws.cell(row=row_num, column=4, value=f"{row['TOTAL%']}%")
        ws.cell(row=row_num, column=5, value=row['CAT1'])
        ws.cell(row=row_num, column=6, value=row['CAT1*'])
        ws.cell(row=row_num, column=7, value=row['CAT2'])
    
    # Add totals row
    total_row = len(result) + 11
    ws.cell(row=total_row, column=3, value=int(result['BOXES'].sum()))
    ws.cell(row=total_row, column=4, value="100.00%")
    ws.cell(row=total_row, column=5, value=int(result['CAT1'].sum()))
    ws.cell(row=total_row, column=6, value=int(result['CAT1*'].sum()))
    ws.cell(row=total_row, column=7, value=int(result['CAT2'].sum()))
    
    # Add borders to the table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(10, total_row + 1):
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = thin_border
    
    # Adjust column widths
    for col in range(2, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    # Create pie chart for size distribution
    pie = PieChart()
    pie.title = "Distribution by Size"
    
    # Define data references - boxes data by size (excluding totals row)
    data_rows = len(result)
    labels = Reference(ws, min_col=2, min_row=11, max_row=10+data_rows)  # SIZE column
    data = Reference(ws, min_col=3, min_row=11, max_row=10+data_rows)    # BOXES column without header
    
    # Add data to chart
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    
    # Remove series title
    pie.series[0].title = None
    
    # Show data labels with values only
    pie.dataLabels = openpyxl.chart.label.DataLabelList()
    pie.dataLabels.showVal = True
    pie.dataLabels.showPercent = False
    pie.dataLabels.showCatName = False
    pie.dataLabels.showLegendKey = False
    pie.dataLabels.showSerName = False
    
    # Position the chart
    pie.width = 12
    pie.height = 12
    ws.add_chart(pie, "B23")
    
    # Create second sheet - breakdown by LOTE
    ws2 = wb.create_sheet("Breakdown by LOTE")
    
    # Group by LOTE, CAT and CALIBRE
    lote_breakdown = df.groupby(['LOTE', 'CAT', 'CALIBRE'])['CANTIDAD - Cajas'].sum().reset_index()
    lote_breakdown = lote_breakdown.sort_values(['LOTE', 'CAT', 'CALIBRE'])
    lote_breakdown.columns = ['LOTE', 'CAT', 'CALIBRE', 'CANTIDAD - CAJAS']
    
    # Add title
    ws2.merge_cells('A1:D1')
    ws2['A1'] = 'BREAKDOWN BY LOTE'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ['LOTE', 'CAT', 'CALIBRE', 'CANTIDAD - CAJAS']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add data
    for idx, row in lote_breakdown.iterrows():
        row_num = idx + 4
        ws2.cell(row=row_num, column=1, value=int(row['LOTE']))
        ws2.cell(row=row_num, column=2, value=str(row['CAT']))
        ws2.cell(row=row_num, column=3, value=int(row['CALIBRE']))
        ws2.cell(row=row_num, column=4, value=int(row['CANTIDAD - CAJAS']))
        
        # Add borders
        for col in range(1, 5):
            ws2.cell(row=row_num, column=col).border = thin_border
    
    # Add summary by LOTE
    current_row = len(lote_breakdown) + 6
    ws2.cell(row=current_row, column=1, value="SUMMARY BY LOTE")
    ws2.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    lote_totals = lote_breakdown.groupby('LOTE')['CANTIDAD - CAJAS'].sum().sort_values(ascending=False)
    for lote, total in lote_totals.items():
        ws2.cell(row=current_row, column=1, value=f"Lote {int(lote)}:")
        ws2.cell(row=current_row, column=4, value=int(total))
        ws2.cell(row=current_row, column=4).font = Font(bold=True)
        current_row += 1
    
    # Add grand total
    current_row += 1
    ws2.cell(row=current_row, column=1, value="GRAND TOTAL:")
    ws2.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws2.cell(row=current_row, column=4, value=int(lote_breakdown['CANTIDAD - CAJAS'].sum()))
    ws2.cell(row=current_row, column=4).font = Font(bold=True, size=12)
    
    # Adjust column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 18
    
    # Create third sheet - breakdown by CÓDIGO DEL LUGAR DE PRODUCCIÓN
    ws3 = wb.create_sheet("Breakdown by Location")
    
    # Group by CÓDIGO DEL LUGAR DE PRODUCCIÓN, CAT and CALIBRE
    location_breakdown = df.groupby(['CÓDIGO DEL LUGAR DE PRODUCCIÓN', 'CAT', 'CALIBRE'])['CANTIDAD - Cajas'].sum().reset_index()
    location_breakdown = location_breakdown.sort_values(['CÓDIGO DEL LUGAR DE PRODUCCIÓN', 'CAT', 'CALIBRE'])
    location_breakdown.columns = ['CÓDIGO DEL LUGAR DE PRODUCCIÓN', 'CAT', 'CALIBRE', 'CANTIDAD - CAJAS']
    
    # Add title
    ws3.merge_cells('A1:D1')
    ws3['A1'] = 'BREAKDOWN BY PRODUCTION LOCATION'
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ['CÓDIGO PRODUCCIÓN', 'CAT', 'CALIBRE', 'CANTIDAD - CAJAS']
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add data
    for idx, row in location_breakdown.iterrows():
        row_num = idx + 4
        ws3.cell(row=row_num, column=1, value=row['CÓDIGO DEL LUGAR DE PRODUCCIÓN'])
        ws3.cell(row=row_num, column=2, value=str(row['CAT']))
        ws3.cell(row=row_num, column=3, value=int(row['CALIBRE']))
        ws3.cell(row=row_num, column=4, value=int(row['CANTIDAD - CAJAS']))
        
        # Add borders
        for col in range(1, 5):
            ws3.cell(row=row_num, column=col).border = thin_border
    
    # Add summary by production location
    current_row = len(location_breakdown) + 6
    ws3.cell(row=current_row, column=1, value="SUMMARY BY PRODUCTION LOCATION")
    ws3.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    location_totals = location_breakdown.groupby('CÓDIGO DEL LUGAR DE PRODUCCIÓN')['CANTIDAD - CAJAS'].sum().sort_values(ascending=False)
    for location, total in location_totals.items():
        ws3.cell(row=current_row, column=1, value=location)
        ws3.cell(row=current_row, column=4, value=int(total))
        ws3.cell(row=current_row, column=4).font = Font(bold=True)
        current_row += 1
    
    # Add grand total
    current_row += 1
    ws3.cell(row=current_row, column=1, value="GRAND TOTAL:")
    ws3.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws3.cell(row=current_row, column=4, value=int(location_breakdown['CANTIDAD - CAJAS'].sum()))
    ws3.cell(row=current_row, column=4).font = Font(bold=True, size=12)
    
    # Adjust column widths
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 8
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 18
    
    # Save the file
    wb.save(output_file)
    print(f"Analysis file created: {output_file}")
    
    # Print summary
    print("\nSummary:")
    print(f"Total boxes: {int(result['BOXES'].sum())}")
    print(f"Total CAT1: {int(result['CAT1'].sum())}")
    print(f"Total CAT1*: {int(result['CAT1*'].sum())}")
    print(f"Total CAT2: {int(result['CAT2'].sum())}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_analysis.py <source_file> [output_file]")
        print("Example: python generate_analysis.py 'PACKING LIST 14.xlsx'")
        sys.exit(1)
    
    source_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' not found")
        sys.exit(1)
    
    generate_packing_analysis(source_file, output_file)