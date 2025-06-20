import json
import base64
import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

def lambda_handler(event, context):
    print(f"Event: {json.dumps(event)}")
    
    try:
        # Get file from request
        if 'body' not in event:
            raise ValueError("No body in request")
            
        body = json.loads(event['body'])
        
        if 'file' not in body:
            raise ValueError("No file in request body")
            
        file_content = base64.b64decode(body['file'])
        filename = body.get('filename', 'packing_list.xlsx')
        
        print(f"Processing file: {filename}, size: {len(file_content)} bytes")
        
        # Process the Excel file
        output_buffer = process_excel(io.BytesIO(file_content), filename)
        
        # Return processed file as base64
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({
                'file': base64.b64encode(output_buffer.getvalue()).decode('utf-8'),
                'filename': filename.replace('.xlsx', '_ANALYSIS.xlsx')
            })
        }
    except Exception as e:
        import traceback
        error_message = f"Error: {str(e)}\nTraceback: {traceback.format_exc()}"
        print(error_message)
        
        return {
            'statusCode': 500,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Content-Type': 'application/json'
            },
            'body': json.dumps({
                'error': str(e),
                'type': type(e).__name__
            })
        }

def process_excel(input_buffer, filename):
    # Read the Excel file
    df = pd.read_excel(input_buffer, sheet_name='DATA', header=1)
    
    # Define size order
    size_order = [12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 36, 40, 48, 50, 60, 70, 84]
    
    # Clean and prepare data
    df['CALIBRE'] = pd.to_numeric(df['CALIBRE'], errors='coerce')
    df['CAT'] = df['CAT'].astype(str).str.strip()
    df['CAT'] = df['CAT'].replace({'1*': 'CAT1*', 'II': 'CAT2', '2': 'CAT2'})
    df['CAT'] = 'CAT' + df['CAT'].astype(str)
    df['CANTIDAD - Cajas'] = pd.to_numeric(df['CANTIDAD - Cajas'], errors='coerce').fillna(0)
    
    # Group by size and category
    grouped = df.groupby(['CALIBRE', 'CAT'])['CANTIDAD - Cajas'].sum().reset_index()
    
    # Create summary by size
    size_summary = grouped.groupby('CALIBRE')['CANTIDAD - Cajas'].sum().reset_index()
    size_summary = size_summary.rename(columns={'CANTIDAD - Cajas': 'Total'})
    
    # Calculate percentages
    total_boxes = size_summary['Total'].sum()
    size_summary['Percentage'] = (size_summary['Total'] / total_boxes * 100).round(2)
    
    # Sort by size order
    size_summary['CALIBRE'] = pd.Categorical(size_summary['CALIBRE'], categories=size_order, ordered=True)
    size_summary = size_summary.sort_values('CALIBRE')
    
    # Create pivot for categories
    category_pivot = grouped.pivot(index='CALIBRE', columns='CAT', values='CANTIDAD - Cajas').fillna(0)
    
    # Merge summary with category breakdown
    result = size_summary.merge(category_pivot, left_on='CALIBRE', right_index=True, how='left')
    
    # Reorder columns
    base_columns = ['CALIBRE', 'Total', 'Percentage']
    cat_columns = [col for col in result.columns if col.startswith('CAT')]
    cat_columns.sort()
    result = result[base_columns + cat_columns]
    
    # Create output Excel file
    output_buffer = io.BytesIO()
    writer = pd.ExcelWriter(output_buffer, engine='openpyxl')
    
    # Write main summary
    result.to_excel(writer, sheet_name='Hoja1', index=False, startrow=6)
    
    # Add headers and totals
    workbook = writer.book
    worksheet = writer.sheets['Hoja1']
    
    # Add title
    worksheet.merge_cells('A1:H1')
    worksheet['A1'] = 'ANÁLISIS DE PACKING LIST'
    worksheet['A1'].font = Font(size=16, bold=True)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    # Add filename
    worksheet['A3'] = f'Archivo: {filename}'
    worksheet['A3'].font = Font(italic=True)
    
    # Add date
    from datetime import datetime
    worksheet['A4'] = f'Fecha de análisis: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    
    # Style the data
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply styles to headers
    for cell in worksheet[7]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    # Apply borders to data
    for row in worksheet.iter_rows(min_row=8, max_row=worksheet.max_row, min_col=1, max_col=len(result.columns)):
        for cell in row:
            cell.border = border
            if cell.column == 1:  # CALIBRE column
                cell.alignment = Alignment(horizontal='center')
    
    # Add totals row
    total_row = worksheet.max_row + 1
    worksheet[f'A{total_row}'] = 'TOTAL'
    worksheet[f'A{total_row}'].font = Font(bold=True)
    worksheet[f'B{total_row}'] = f'=SUM(B8:B{total_row-1})'
    worksheet[f'C{total_row}'] = '100.00'
    
    # Sum category columns
    for i, col in enumerate(cat_columns, start=4):
        col_letter = get_column_letter(i)
        worksheet[f'{col_letter}{total_row}'] = f'=SUM({col_letter}8:{col_letter}{total_row-1})'
    
    # Style totals row
    for cell in worksheet[total_row]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.border = border
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            if hasattr(cell, 'column_letter'):
                if column_letter is None:
                    column_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Add pie chart
    pie = PieChart()
    pie.title = "Distribución por Calibre"
    labels = Reference(worksheet, min_col=1, min_row=8, max_row=total_row-1)
    data = Reference(worksheet, min_col=2, min_row=7, max_row=total_row-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 10
    pie.width = 15
    worksheet.add_chart(pie, f"{get_column_letter(len(result.columns)+2)}5")
    
    # Add breakdown sheets
    lote_breakdown = df.groupby(['LOTE', 'CALIBRE', 'CAT'])['CANTIDAD - Cajas'].sum().reset_index()
    lote_breakdown.to_excel(writer, sheet_name='Breakdown by LOTE', index=False)
    
    location_breakdown = df.groupby(['CÓDIGO DEL LUGAR DE PRODUCCIÓN', 'CALIBRE', 'CAT'])['CANTIDAD - Cajas'].sum().reset_index()
    location_breakdown.to_excel(writer, sheet_name='Breakdown by Location', index=False)
    
    writer.close()
    output_buffer.seek(0)
    return output_buffer